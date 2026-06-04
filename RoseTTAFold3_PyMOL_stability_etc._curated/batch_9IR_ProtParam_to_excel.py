# -*- coding: utf-8 -*-

from pathlib import Path
from Bio import SeqIO
from Bio.SeqUtils.ProtParam import ProteinAnalysis
import pandas as pd
import re

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule


# =========================================================
# 1. 输入和输出文件夹
# =========================================================

input_dir = Path(
    r"C:\Users\Lenovo\Desktop\small_molecule_design\5_Filtering_structure\in_9IR"
)

output_dir = Path(
    r"C:\Users\Lenovo\Desktop\small_molecule_design\5_Filtering_structure\out_9IR"
)

output_dir.mkdir(parents=True, exist_ok=True)


# =========================================================
# 2. 标准氨基酸与 N-end rule 半衰期估计表
# =========================================================

valid_amino_acids = set("ACDEFGHIKLMNPQRSTVWY")

half_life_table = {
    "A": {"Mammalian": "4.4 h",  "E_coli": ">10 h"},
    "R": {"Mammalian": "1 h",    "E_coli": "2 min"},
    "N": {"Mammalian": "1.4 h",  "E_coli": ">10 h"},
    "D": {"Mammalian": "1.1 h",  "E_coli": ">10 h"},
    "C": {"Mammalian": "1.2 h",  "E_coli": ">10 h"},
    "Q": {"Mammalian": "0.8 h",  "E_coli": ">10 h"},
    "E": {"Mammalian": "1 h",    "E_coli": ">10 h"},
    "G": {"Mammalian": "30 h",   "E_coli": ">10 h"},
    "H": {"Mammalian": "3.5 h",  "E_coli": ">10 h"},
    "I": {"Mammalian": "20 h",   "E_coli": ">10 h"},
    "L": {"Mammalian": "5.5 h",  "E_coli": "2 min"},
    "K": {"Mammalian": "1.3 h",  "E_coli": "2 min"},
    "M": {"Mammalian": "30 h",   "E_coli": ">10 h"},
    "F": {"Mammalian": "1.1 h",  "E_coli": "2 min"},
    "P": {"Mammalian": ">20 h",  "E_coli": "Unknown"},
    "S": {"Mammalian": "1.9 h",  "E_coli": ">10 h"},
    "T": {"Mammalian": "7.2 h",  "E_coli": ">10 h"},
    "W": {"Mammalian": "2.8 h",  "E_coli": "2 min"},
    "Y": {"Mammalian": "2.8 h",  "E_coli": "2 min"},
    "V": {"Mammalian": "100 h",  "E_coli": ">10 h"},
}


# =========================================================
# 3. 从 FASTA 标题中提取信息，并简化候选序列编号
# =========================================================

def extract_header_information(description: str):
    """
    输入示例：
    partial_0_model_0_b0_d0, sequence_recovery=0.5630,
    ligand_interface_sequence_recovery=0.6111

    输出 candidate_id：
    partial_0_model_0_0
    """

    original_id = description.split(",")[0].strip()

    # 将 _b0_d0、_b0_d1 等形式转换为 _0、_1
    candidate_id = re.sub(
        r"_b\d+_d(\d+)$",
        r"_\1",
        original_id
    )

    recovery_match = re.search(
        r"(?:^|,\s*)sequence_recovery=([0-9]*\.?[0-9]+)(?=,|$)",
        description
    )

    interface_match = re.search(
        r"(?:^|,\s*)ligand_interface_sequence_recovery=([0-9]*\.?[0-9]+)(?=,|$)",
        description
    )

    sequence_recovery = (
        float(recovery_match.group(1))
        if recovery_match else None
    )

    ligand_interface_sequence_recovery = (
        float(interface_match.group(1))
        if interface_match else None
    )

    return (
        candidate_id,
        sequence_recovery,
        ligand_interface_sequence_recovery
    )


# =========================================================
# 4. 文件名自然排序函数
# =========================================================

def natural_sort_key(path: Path):
    """
    确保 model_2 排在 model_10 前面。
    """

    parts = re.split(r"(\d+)", path.name)

    return [
        int(part) if part.isdigit() else part.lower()
        for part in parts
    ]


# =========================================================
# 5. 分析单个 FASTA 文件
# =========================================================

def analyze_one_fasta(fasta_file: Path):

    results = []
    skipped_sequences = []

    for record in SeqIO.parse(str(fasta_file), "fasta"):

        sequence = re.sub(r"\s+", "", str(record.seq)).upper()

        (
            candidate_id,
            sequence_recovery,
            ligand_interface_sequence_recovery
        ) = extract_header_information(record.description)

        if not sequence:
            skipped_sequences.append(
                f"{candidate_id}: 序列为空"
            )
            continue

        invalid_residues = sorted(
            set(sequence) - valid_amino_acids
        )

        if invalid_residues:
            skipped_sequences.append(
                f"{candidate_id}: 含有非标准氨基酸 "
                f"{','.join(invalid_residues)}"
            )
            continue

        analysis = ProteinAnalysis(sequence)

        instability_index = round(
            analysis.instability_index(),
            2
        )

        if instability_index < 40:
            stability_prediction = "Stable"
        else:
            stability_prediction = "Potentially unstable"

        n_terminal_residue = sequence[0]
        half_life = half_life_table[n_terminal_residue]

        results.append({
            "candidate_id": candidate_id,
            "sequence_recovery": sequence_recovery,
            "ligand_interface_sequence_recovery": ligand_interface_sequence_recovery,
            "instability_index": instability_index,
            "stability_prediction": stability_prediction,
            "half_life_mammalian_reticulocytes_in_vitro": half_life["Mammalian"],
            "half_life_E_coli_in_vivo": half_life["E_coli"]
        })

    return results, skipped_sequences


# =========================================================
# 6. 将单个 FASTA 的结果保存为 Excel
# =========================================================

def write_result_to_excel(results, output_file: Path):

    columns = [
        "candidate_id",
        "sequence_recovery",
        "ligand_interface_sequence_recovery",
        "instability_index",
        "stability_prediction",
        "half_life_mammalian_reticulocytes_in_vitro",
        "half_life_E_coli_in_vivo"
    ]

    df = pd.DataFrame(results, columns=columns)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        df.to_excel(
            writer,
            sheet_name="ProtParam_results",
            index=False
        )

        worksheet = writer.book["ProtParam_results"]

        # 固定表头
        worksheet.freeze_panes = "A2"

        # 开启筛选
        worksheet.auto_filter.ref = worksheet.dimensions

        # 表头格式
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # 数据格式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

        # 数值显示格式
        for row_number in range(2, worksheet.max_row + 1):
            worksheet[f"B{row_number}"].number_format = "0.0000"
            worksheet[f"C{row_number}"].number_format = "0.0000"
            worksheet[f"D{row_number}"].number_format = "0.00"

        # 列宽
        column_widths = {
            "A": 32,
            "B": 20,
            "C": 36,
            "D": 18,
            "E": 24,
            "F": 44,
            "G": 30
        }

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

        worksheet.row_dimensions[1].height = 34

        # instability_index 条件格式
        # < 40 显示绿色；>= 40 显示浅红色
        if worksheet.max_row >= 2:

            green_fill = PatternFill(
                fill_type="solid",
                fgColor="E2F0D9"
            )

            red_fill = PatternFill(
                fill_type="solid",
                fgColor="FCE4D6"
            )

            instability_range = f"D2:D{worksheet.max_row}"

            worksheet.conditional_formatting.add(
                instability_range,
                CellIsRule(
                    operator="lessThan",
                    formula=["40"],
                    fill=green_fill
                )
            )

            worksheet.conditional_formatting.add(
                instability_range,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["40"],
                    fill=red_fill
                )
            )


# =========================================================
# 7. 批量处理全部 FASTA 文件
# =========================================================

def main():

    fasta_files = sorted(
        input_dir.glob("*.fa"),
        key=natural_sort_key
    )

    if not fasta_files:
        raise FileNotFoundError(
            f"在输入文件夹中没有找到 .fa 文件：{input_dir}"
        )

    print("\n====================================================")
    print("9IR ProtParam 批量分析开始")
    print("====================================================")
    print(f"输入文件夹：{input_dir}")
    print(f"输出文件夹：{output_dir}")
    print(f"检测到 FASTA 文件数：{len(fasta_files)}")

    if len(fasta_files) != 75:
        print(
            f"[提醒] 当前检测到 {len(fasta_files)} 个 FASTA 文件，"
            "与预期的 75 个不一致，请检查输入文件夹。"
        )

    total_sequences = 0
    total_skipped = 0
    generated_files = 0

    for fasta_file in fasta_files:

        results, skipped_sequences = analyze_one_fasta(fasta_file)

        output_file = output_dir / (
            fasta_file.stem + "_ProtParam_results.xlsx"
        )

        write_result_to_excel(results, output_file)

        total_sequences += len(results)
        total_skipped += len(skipped_sequences)
        generated_files += 1

        stable_number = sum(
            1 for item in results
            if item["stability_prediction"] == "Stable"
        )

        unstable_number = len(results) - stable_number

        print("\n----------------------------------------------------")
        print(f"输入文件：{fasta_file.name}")
        print(f"成功分析序列数：{len(results)}")
        print(f"预测稳定序列数：{stable_number}")
        print(f"预测可能不稳定序列数：{unstable_number}")
        print(f"输出文件：{output_file.name}")

        if skipped_sequences:
            print("跳过的序列：")
            for message in skipped_sequences:
                print(f"  - {message}")

    print("\n====================================================")
    print("9IR ProtParam 批量分析完成")
    print("====================================================")
    print(f"共处理 FASTA 文件数：{len(fasta_files)}")
    print(f"共生成 Excel 文件数：{generated_files}")
    print(f"共成功分析序列数：{total_sequences}")
    print(f"共跳过序列数：{total_skipped}")
    print(f"结果输出位置：{output_dir}")


if __name__ == "__main__":
    main()